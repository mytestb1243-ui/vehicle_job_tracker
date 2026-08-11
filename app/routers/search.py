import datetime
import io
import urllib.parse

from fastapi import APIRouter, Request, Depends
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.models import Job
from app.choices import (
    CHECKLIST_FIELDS,
    DEVICE_TYPE_OPTIONS,
    COMPANY_NAME_OPTIONS,
    SERVICE_TYPE_OPTIONS,
    COLUMN_ORDER,
)
from app.dependencies import get_current_user, get_flashed_messages
from app.render import templates

router = APIRouter()

ALLOWED_PAGE_SIZES = [10, 25, 50, 100, 200]

CHECKED = "\u2611"    # ☑
UNCHECKED = "\u2610"  # ☐


def _apply_filters(query, vehicle_or_device, company_name, location, device_type,
                    service_type, date_from, date_to):
    if vehicle_or_device:
        like = f"%{vehicle_or_device.strip()}%"
        query = query.filter(
            or_(Job.vehicle_no.ilike(like), Job.device_id.ilike(like))
        )
    if company_name:
        query = query.filter(Job.company_name == company_name)
    if location:
        query = query.filter(Job.location.ilike(f"%{location.strip()}%"))
    if device_type:
        query = query.filter(Job.device_type == device_type)
    if service_type:
        query = query.filter(Job.service_type == service_type)
    if date_from:
        try:
            df = datetime.datetime.strptime(date_from, "%Y-%m-%d").date()
            query = query.filter(Job.job_date >= df)
        except ValueError:
            pass
    if date_to:
        try:
            dt_ = datetime.datetime.strptime(date_to, "%Y-%m-%d").date()
            query = query.filter(Job.job_date <= dt_)
        except ValueError:
            pass
    return query


@router.get("/search")
def search_page(
    request: Request,
    vehicle_or_device: str = "",
    company_name: str = "",
    location: str = "",
    device_type: str = "",
    service_type: str = "",
    date_from: str = "",
    date_to: str = "",
    page: int = 1,
    page_size: int = 10,
    db: Session = Depends(get_db),
):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    if page_size not in ALLOWED_PAGE_SIZES:
        page_size = 10
    page = max(page, 1)

    query = db.query(Job)
    query = _apply_filters(
        query, vehicle_or_device, company_name, location, device_type,
        service_type, date_from, date_to,
    )
    query = query.order_by(Job.job_date.desc(), Job.id.desc())

    total_results = query.count()
    total_pages = max((total_results + page_size - 1) // page_size, 1)
    page = min(page, total_pages)

    results = query.offset((page - 1) * page_size).limit(page_size).all()

    filters_qs = urllib.parse.urlencode(
        {
            "vehicle_or_device": vehicle_or_device,
            "company_name": company_name,
            "location": location,
            "device_type": device_type,
            "service_type": service_type,
            "date_from": date_from,
            "date_to": date_to,
        }
    )

    return templates.TemplateResponse(
        request,
        "search.html",
        {
            "user": user,
            "messages": get_flashed_messages(request),
            "jobs": results,
            "total_results": total_results,
            "page": page,
            "page_size": page_size,
            "page_sizes": ALLOWED_PAGE_SIZES,
            "total_pages": total_pages,
            "device_types": DEVICE_TYPE_OPTIONS,
            "company_names": COMPANY_NAME_OPTIONS,
            "service_types": SERVICE_TYPE_OPTIONS,
            "checklist_fields": CHECKLIST_FIELDS,
            "filters_qs": filters_qs,
            "f_vehicle_or_device": vehicle_or_device,
            "f_company_name": company_name,
            "f_location": location,
            "f_device_type": device_type,
            "f_service_type": service_type,
            "f_date_from": date_from,
            "f_date_to": date_to,
            "active_page": "search",
        },
    )


@router.get("/search/export")
def export_excel(
    request: Request,
    vehicle_or_device: str = "",
    company_name: str = "",
    location: str = "",
    device_type: str = "",
    service_type: str = "",
    date_from: str = "",
    date_to: str = "",
    db: Session = Depends(get_db),
):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    query = db.query(Job)
    query = _apply_filters(
        query, vehicle_or_device, company_name, location, device_type,
        service_type, date_from, date_to,
    )
    jobs = query.order_by(Job.job_date.asc(), Job.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    # Headers exactly matching column_information.txt, checklist expanded.
    headers = (
        COLUMN_ORDER[:13]
        + [label for _name, label in CHECKLIST_FIELDS]
        + ["Notes"]
    )

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    link_font = Font(color="1155CC", underline="single")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    def _link_cell(row, col, value, label="CLICK HERE"):
        cell = ws.cell(row=row, column=col, value=label if value else "")
        if value:
            link = value
            if not (link.startswith("http://") or link.startswith("https://")):
                link = "https://" + link
            cell.hyperlink = link
            cell.font = link_font
        cell.alignment = center
        return cell

    row_idx = 2
    for job in jobs:
        col = 1
        ws.cell(row=row_idx, column=col, value=job.vehicle_no).alignment = center
        col += 1
        ws.cell(row=row_idx, column=col, value=job.device_id or "").alignment = center
        col += 1
        ws.cell(row=row_idx, column=col, value=job.company_name or "").alignment = left
        col += 1
        date_str = f"{job.job_date.month}/{job.job_date.day}/{job.job_date.year}" if job.job_date else ""
        ws.cell(row=row_idx, column=col, value=date_str).alignment = center
        col += 1
        ws.cell(row=row_idx, column=col, value=job.location).alignment = left
        col += 1
        _link_cell(row_idx, col, job.before_images)
        col += 1
        _link_cell(row_idx, col, job.after_images)
        col += 1
        _link_cell(row_idx, col, job.service_form)
        col += 1
        ws.cell(row=row_idx, column=col, value=job.po_number or "").alignment = center
        col += 1
        ws.cell(row=row_idx, column=col, value=job.device_type or "").alignment = center
        col += 1
        ws.cell(row=row_idx, column=col, value=job.service_type or "").alignment = center
        col += 1
        ws.cell(row=row_idx, column=col, value=job.tempering or "-").alignment = left
        col += 1
        _link_cell(row_idx, col, job.tempering_evidence)
        col += 1

        for field_name, _label in CHECKLIST_FIELDS:
            checked = getattr(job, field_name)
            c = ws.cell(row=row_idx, column=col, value=CHECKED if checked else UNCHECKED)
            c.alignment = center
            c.font = Font(size=13)
            col += 1

        ws.cell(row=row_idx, column=col, value=job.notes or "").alignment = left

        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=c_idx).border = thin_border

        row_idx += 1

    widths = [12, 12, 18, 11, 16, 12, 12, 12, 12, 11, 16, 16, 14] + [11] * len(CHECKLIST_FIELDS) + [40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 34

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"vehicle_job_tracker_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
